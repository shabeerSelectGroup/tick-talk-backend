"""Excel exports for participants, matches, and leaderboard."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def _style_header(ws, row: int = 1) -> None:
    fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    bold = Font(bold=True, color="FFFFFF")
    for cell in ws[row]:
        cell.fill = fill
        cell.font = bold


def _autosize(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 2, 48)


def _write_sheet(ws, headers: list[str], rows: list[list]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)
    _style_header(ws)
    _autosize(ws)


def generate_excel_participants(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Participants"
    headers = [
        "ID",
        "Name",
        "Email",
        "Company",
        "Title",
        "Score",
        "Rank",
        "Tasks completed",
        "Matches",
        "Progress %",
        "Joined at",
    ]
    data = [
        [
            r["id"],
            r["display_name"],
            r["email"],
            r["company"],
            r["title"],
            r["score"],
            r["rank"],
            r["tasks_completed"],
            r["matches_count"],
            r["progress_percent"],
            r["joined_at"],
        ]
        for r in rows
    ]
    _write_sheet(ws, headers, data)
    return _workbook_bytes(wb)


def generate_excel_matches(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Matches"
    headers = [
        "ID",
        "Initiator",
        "Initiator company",
        "Partner",
        "Partner company",
        "Task",
        "Type",
        "Points",
        "Created at",
    ]
    data = [
        [
            r["id"],
            r["initiator_name"],
            r["initiator_company"],
            r["partner_name"],
            r["partner_company"],
            r["task_title"],
            r["match_type"],
            r["points_awarded"],
            r["created_at"],
        ]
        for r in rows
    ]
    _write_sheet(ws, headers, data)
    return _workbook_bytes(wb)


def generate_excel_leaderboard(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Leaderboard"
    headers = ["Rank", "Participant ID", "Name", "Company", "Score", "Tasks", "Finished at"]
    data = [
        [
            r.get("rank", ""),
            r.get("participant_id", ""),
            r.get("display_name", ""),
            r.get("company") or "",
            r.get("score", 0),
            r.get("tasks_completed", 0),
            r.get("finished_at") or "",
        ]
        for r in rows
    ]
    _write_sheet(ws, headers, data)
    return _workbook_bytes(wb)


def generate_excel_bundle(
    participants: list[dict],
    matches: list[dict],
    leaderboard: list[dict],
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("Participants")
    headers_p = ["ID", "Name", "Email", "Company", "Score", "Rank", "Tasks", "Matches", "Joined"]
    ws1.append(headers_p)
    for r in participants:
        ws1.append(
            [
                r["id"],
                r["display_name"],
                r["email"],
                r["company"],
                r["score"],
                r["rank"],
                r["tasks_completed"],
                r["matches_count"],
                r["joined_at"],
            ]
        )
    _style_header(ws1)
    _autosize(ws1)

    ws2 = wb.create_sheet("Matches")
    headers_m = ["ID", "Initiator", "Partner", "Task", "Type", "Points", "Created"]
    ws2.append(headers_m)
    for r in matches:
        ws2.append(
            [
                r["id"],
                r["initiator_name"],
                r["partner_name"],
                r["task_title"],
                r["match_type"],
                r["points_awarded"],
                r["created_at"],
            ]
        )
    _style_header(ws2)
    _autosize(ws2)

    ws3 = wb.create_sheet("Leaderboard")
    headers_l = ["Rank", "Name", "Company", "Score", "Tasks", "Finished"]
    ws3.append(headers_l)
    for r in leaderboard:
        ws3.append(
            [
                r.get("rank"),
                r.get("display_name"),
                r.get("company") or "",
                r.get("score"),
                r.get("tasks_completed"),
                r.get("finished_at") or "",
            ]
        )
    _style_header(ws3)
    _autosize(ws3)

    return _workbook_bytes(wb)


def _workbook_bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
